#pip3 install openpyxl
#pip install mysql-connector-python
#pip3 install passlib
import mysql.connector
mydb = mysql.connector.connect(
  host="127.0.0.1",
  user="dmoj",
  password="xxxxxxxx",
  database="dmoj"
)

import openpyxl
wb_obj = openpyxl.load_workbook('taikhoan.xlsx') 
# Read the active sheet:
sheet = wb_obj.active
tatcadong = []
for row in sheet.iter_rows():
    dong = []
    for j in range(sheet.max_column):
        dong.append(row[j].value)
    tatcadong.append(dong)
tatcadong.pop(0)

import base64
import hashlib
import secrets
ALGORITHM = "pbkdf2_sha256"


def hash_password(password, salt=None, iterations=260000):
    if salt is None:
        salt = secrets.token_hex(16)
    assert salt and isinstance(salt, str) and "$" not in salt
    assert isinstance(password, str)
    pw_hash = hashlib.pbkdf2_hmac("sha256", password.encode("utf-8"), salt.encode("utf-8"), iterations)
    b64_hash = base64.b64encode(pw_hash).decode("ascii").strip()
    return "{}${}${}${}".format(ALGORITHM, iterations, salt, b64_hash)


for rx in tatcadong:
    id, hoten, bang, taikhoan, mk = rx[0], rx[1], rx[2],rx[3], rx[4]
    mk = hash_password(mk, None, 260000)
    mycursor = mydb.cursor()
    sql = "INSERT INTO auth_user VALUES (%s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s)"
    val = (id+20, mk, '2023-4-4 4:4:4', '0', taikhoan, hoten, bang, taikhoan+'@gmail.com', '0', '1', '2023-1-1')
    mycursor.execute(sql, val)
    mydb.commit()



print('Them xong! Nho khoi dong lai may chu de DANG NHAP')
